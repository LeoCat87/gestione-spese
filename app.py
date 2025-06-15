import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import gdown
import os
st.set_page_config(page_title="Gestione Spese", layout="wide")
# === CONFIGURAZIONE ===
GDRIVE_FILE_ID = "1PJ9TCcq4iBHeg8CpC1KWss0UWSg86BJn"
EXCEL_PATH = "Spese_App.xlsx"
# Scarica il file Excel da Google Drive
@st.cache_data
def scarica_excel_da_drive():
    if not os.path.exists(EXCEL_PATH):  # Scarica solo se non esiste già
        url = f"https://drive.google.com/uc?id={GDRIVE_FILE_ID}"
        gdown.download(url, EXCEL_PATH, quiet=True)
scarica_excel_da_drive()
# === FUNZIONI DI CARICAMENTO ===
@st.cache_data
def carica_spese():
    sheet = pd.read_excel(EXCEL_PATH, sheet_name="Spese Leo", header=None)
    mesi_excel = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                  "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

    col_mese = {}
    for col_idx in range(sheet.shape[1]):
        cella = sheet.iloc[0, col_idx]
        if isinstance(cella, str) and cella.lower() in mesi_excel:
            col_mese[cella.lower()] = col_idx

    spese = []
    for mese_lower, start_col in col_mese.items():
        intestazioni = sheet.iloc[1, start_col:start_col+3].tolist()
        if "Valore" in intestazioni and "Tag" in intestazioni:
            df_blocco = sheet.iloc[2:, start_col:start_col+3].copy()
            df_blocco.columns = intestazioni
            df_blocco["Mese"] = mese_lower.capitalize()
            spese.append(df_blocco)

    if spese:
        df = pd.concat(spese, ignore_index=True)
        df = df.dropna(subset=["Valore", "Tag"])
        df["Valore"] = pd.to_numeric(df["Valore"], errors="coerce").fillna(0)
        df["Testo"] = df.get("Testo", "").fillna("")

        def categoria_per_tag(tag):
            if tag in ["Stipendio", "Entrate extra"]:
                return "Entrate"
            elif tag in ["Affitto", "Bollette", "Spesa", "Abbonamenti", "Trasporti", "Assicurazione"]:
                return "Uscite necessarie"
            else:
                return "Uscite variabili"

        df["Categoria"] = df["Tag"].apply(categoria_per_tag)
        return df
    else:
        return pd.DataFrame(columns=["Testo", "Valore", "Tag", "Mese", "Categoria"])

@st.cache_data
def carica_riepilogo():
    df = pd.read_excel(EXCEL_PATH, sheet_name="Riepilogo Leo", index_col=0)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    return df
@st.cache_data
def carica_dashboard():
    df = pd.read_excel(EXCEL_PATH, sheet_name="Riepilogo Leo", index_col=0)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df["Total"] = df.get("Total", pd.Series(0))  # Se manca "Total", metti 0
    return df
def formatta_euro(val):
    return f"€ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
# === INTERFACCIA ===
st.sidebar.title("📁 Navigazione")
vista = st.sidebar.radio("Scegli una vista:", ["Spese dettagliate", "Riepilogo mensile", "Dashboard"])

# === VISTA 1: SPESE DETTAGLIATE ===
if vista == "Spese dettagliate":
    st.title("📌 Spese Dettagliate")
    df_spese = carica_spese()
    df_riepilogo = carica_riepilogo()

    mesi_disponibili = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                        "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    col1, col2 = st.columns([1, 5])  # col1 = vuota o futura, col2 = filtro + tabella
    with col2:
        mese_sel = st.selectbox("Seleziona il mese:", mesi_disponibili)

    categorie_tag = [str(tag) for tag in df_riepilogo.index if pd.notnull(tag)]
    df_filtrato = df_spese[df_spese["Mese"] == mese_sel].copy()

    if df_filtrato.empty:
        st.info("🔍 Nessuna spesa registrata per il mese selezionato.")
    else:
        with col2:
            edited_df = st.data_editor(
                df_filtrato[["Testo", "Valore", "Tag"]],
                use_container_width=False,
                hide_index=True,
                column_config={
                    "Testo": st.column_config.TextColumn(
                        label="Descrizione",
                        help="Testo libero"
                    ),
                    "Valore": st.column_config.NumberColumn(
                        label="€",
                        help="Importo della spesa",
                        step=0.01,
                        format="€ %.2f"
                    ),
                    "Tag": st.column_config.SelectboxColumn(
                        label="Tag",
                        help="Categoria",
                        options=categorie_tag,
                        required=True
                    )
                }
            )

        if not edited_df.equals(df_filtrato[["Testo", "Valore", "Tag"]]):
            with col2:
                st.success("✅ Modifiche rilevate.")
                if st.button("💾 Salva modifiche"):
                    df_aggiornato = df_spese[df_spese["Mese"] != mese_sel].copy()
                    edited_df["Mese"] = mese_sel
                    edited_df["Valore"] = pd.to_numeric(edited_df["Valore"], errors="coerce").fillna(0)

                    def categoria_per_tag(tag):
                        if tag in ["Stipendio", "Entrate extra"]:
                            return "Entrate"
                        elif tag in ["Affitto", "Bollette", "Spesa", "Abbonamenti", "Trasporti", "Assicurazione"]:
                            return "Uscite necessarie"
                        else:
                            return "Uscite variabili"

                    edited_df["Categoria"] = edited_df["Tag"].apply(categoria_per_tag)
                    edited_df["Testo"] = edited_df["Testo"].fillna("")

                    df_finale = pd.concat([df_aggiornato, edited_df], ignore_index=True)

                    import openpyxl
                    wb = openpyxl.load_workbook(EXCEL_PATH)
                    ws = wb["Spese Leo"]

                    mesi_excel = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                                  "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]
                    mese_col_start = None
                    for col in range(1, ws.max_column + 1):
                        val = ws.cell(row=1, column=col).value
                        if val and isinstance(val, str) and val.lower() == mese_sel.lower():
                            mese_col_start = col
                            break

                    if mese_col_start:
                        for row in range(3, ws.max_row + 1):
                            for c in range(mese_col_start, mese_col_start + 3):
                                ws.cell(row=row, column=c).value = None

                        ws.cell(row=2, column=mese_col_start).value = "Testo"
                        ws.cell(row=2, column=mese_col_start + 1).value = "Valore"
                        ws.cell(row=2, column=mese_col_start + 2).value = "Tag"

                        for i, row in edited_df.iterrows():
                            ws.cell(row=3 + i, column=mese_col_start).value = row["Testo"]
                            ws.cell(row=3 + i, column=mese_col_start + 1).value = float(row["Valore"])
                            ws.cell(row=3 + i, column=mese_col_start + 2).value = row["Tag"]

                        wb.save(EXCEL_PATH)
                        st.success("✅ Modifiche salvate correttamente.")
                    else:
                        st.error("❌ Colonna del mese non trovata nel foglio Excel.")
with open(EXCEL_PATH, "rb") as f:
    st.download_button(
        label="📥 Scarica file aggiornato",
        data=f,
        file_name="Spese_App_aggiornato.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === VISTA 2: RIEPILOGO MENSILE ===
elif vista == "Riepilogo mensile":
    st.title("📊 Riepilogo Mensile per Tag")

    mappa_macrocategorie = {
        "📌 Entrate": ["Stipendio", "Affitto Savoldo 4 + generico"],
        "📌 Uscite necessarie": [
            "PAC Investimenti", "Donazioni (StC, Unicef, Greenpeace)", "Mutuo", "Luce&Gas",
            "Internet/Telefono", "Mezzi", "Spese condominiali", "Spese comuni",
            "Auto (benzina, noleggio, pedaggi, parcheggi)", "Spesa cibo", "Tari", "Unobravo"
        ],
        "📌 Uscite variabili": [
            "Amazon", "Bolli governativi", "Farmacia/Visite", "Food Delivery", "Generiche", "Multa",
            "Uscite (Pranzi,Cena,Apericena,Pub,etc)", "Prelievi", "Regali", "Sharing (auto, motorino, bici)",
            "Shopping (vestiti, mobili,...)", "Stireria", "Viaggi (treno, aereo, hotel, attrazioni, concerti, cinema)"
        ]
    }

    # Carica foglio
    sheet = pd.read_excel(EXCEL_PATH, sheet_name="Spese Leo", header=None)
    mesi_excel = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                  "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

    # Trova blocchi per mese
    col_mese = {}
    for col_idx in range(sheet.shape[1]):
        cella = sheet.iloc[0, col_idx]
        if isinstance(cella, str) and cella.lower() in mesi_excel:
            col_mese[cella.lower()] = col_idx

    spese_totali = []
    for mese_lower, start_col in col_mese.items():
        intestazioni = sheet.iloc[1, start_col:start_col+3].tolist()
        if "Valore" in intestazioni and "Tag" in intestazioni:
            df_blocco = sheet.iloc[2:, start_col:start_col+3].copy()
            df_blocco.columns = intestazioni
            df_blocco = df_blocco.dropna(subset=["Valore", "Tag"])
            df_blocco["Mese"] = mese_lower.capitalize()
            spese_totali.append(df_blocco)

    if spese_totali:
        df_spese = pd.concat(spese_totali, ignore_index=True)
        df_riepilogo = df_spese.groupby(["Tag", "Mese"])["Valore"].sum().unstack(fill_value=0)

        mesi_ordinati = [m.capitalize() for m in mesi_excel]
        df_riepilogo = df_riepilogo.reindex(columns=mesi_ordinati, fill_value=0)

        # Calcola "Media YTD" fino al mese precedente
        from datetime import datetime
        mese_corr = datetime.today().month
        mesi_da_media = mesi_ordinati[:mese_corr - 1] if mese_corr > 1 else []
        if mesi_da_media:
            df_riepilogo["Media YTD"] = df_riepilogo[mesi_da_media].mean(axis=1)
        else:
            df_riepilogo["Media YTD"] = 0

        # Costruzione tabella HTML
        html = """
        <style>
        table {border-collapse: collapse; width: auto; table-layout: auto;}
        th, td {padding: 6px 12px; text-align: left; white-space: nowrap;}
        .macro {background-color: #f0f0f0; font-weight: bold;}
        </style>
        <table>
            <tr>
                <th>Categoria</th>""" + "".join(f"<th>{mese}</th>" for mese in mesi_ordinati) + "<th>Media YTD</th></tr>"

        for macro, tags in mappa_macrocategorie.items():
            html += f'<tr class="macro"><td colspan="{len(mesi_ordinati)+2}">{macro}</td></tr>'
            for tag in tags:
                if tag in df_riepilogo.index:
                    r = df_riepilogo.loc[tag]
                    html += f"<tr><td>{tag}</td>"
                    for mese in mesi_ordinati:
                        euro = formatta_euro(r[mese]) if r[mese] else "€ 0,00"
                        html += f"<td>{euro}</td>"
                    media = formatta_euro(r["Media YTD"]) if r["Media YTD"] else "€ 0,00"
                    html += f"<td>{media}</td></tr>"

        html += "</table>"
        st.markdown(html, unsafe_allow_html=True)
    else:
        st.warning("Nessuna spesa trovata nei blocchi mensili del foglio 'Spese Leo'.")

# === VISTA 3: DASHBOARD ===
elif vista == "Dashboard":
    st.title("📈 Dashboard")

    df_riepilogo = carica_riepilogo()

    mappa_macrocategorie = {
        "Entrate": ["Stipendio", "Affitto Savoldo 4 + generico"],
        "Uscite necessarie": [
            "PAC Investimenti", "Donazioni (StC, Unicef, Greenpeace)", "Mutuo", "Luce&Gas",
            "Internet/Telefono", "Mezzi", "Spese condominiali", "Spese comuni",
            "Auto (benzina, noleggio, pedaggi, parcheggi)", "Spesa cibo", "Tari", "Unobravo"
        ],
        "Uscite variabili": [
            "Amazon", "Bolli governativi", "Farmacia/Visite", "Food Delivery", "Generiche", "Multa",
            "Uscite (Pranzi,Cena,Apericena,Pub,etc)", "Prelievi", "Regali", "Sharing (auto, motorino, bici)",
            "Shopping (vestiti, mobili,...)", "Stireria", "Viaggi (treno, aereo, hotel, attrazioni, concerti, cinema)"
        ]
    }

    mesi = df_riepilogo.columns.tolist()
    df_macrocategorie = pd.DataFrame(columns=mesi)

    for macro, sottotag in mappa_macrocategorie.items():
        tag_presenti = [t for t in sottotag if t in df_riepilogo.index]
        somma = df_riepilogo.loc[tag_presenti].sum() if tag_presenti else pd.Series([0]*len(mesi), index=mesi)
        df_macrocategorie.loc[macro] = somma

    df_macrocategorie.loc["Risparmio mese"] = (
        df_macrocategorie.loc["Entrate"]
        - df_macrocategorie.loc["Uscite necessarie"]
        - df_macrocategorie.loc["Uscite variabili"]
    )
    df_macrocategorie.loc["Risparmio cumulato"] = df_macrocategorie.loc["Risparmio mese"].cumsum()

    # Calcola Media YTD fino al mese precedente
    from datetime import datetime
    mese_corr = datetime.today().month
    mesi_ytd = mesi[:mese_corr - 1] if mese_corr > 1 else []
    df_macrocategorie["Media YTD"] = df_macrocategorie[mesi_ytd].mean(axis=1) if mesi_ytd else 0

    # Tabella formattata
    df_tabella = df_macrocategorie.copy().reset_index().rename(columns={"index": "Voce"})
    for col in df_tabella.columns[1:]:
        df_tabella[col] = df_tabella[col].apply(lambda x: formatta_euro(x) if pd.notnull(x) else "€ 0,00")

    st.subheader("📊 Tabella riepilogo")
    st.dataframe(df_tabella, hide_index=True)

    df_grafico = df_macrocategorie[mesi].transpose()
    st.subheader("📈 Andamento mensile")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(12, 6))
    df_grafico.plot(kind='bar', ax=ax)
    ax.set_title("Entrate, Uscite e Risparmio per mese")
    ax.set_xlabel("Mese")
    ax.set_ylabel("Importo (€)")
    plt.xticks(rotation=45)
    st.pyplot(fig)
